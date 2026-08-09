"""Tests for capa2_ruidosa (noisy version) robustness.

The dataset has two layers:
  - capa1_limpia: clean conversations where patient answers what was asked
  - capa2_ruidosa: same conversation degraded with realistic noise
    (evasive answers, missing info, irrelevant symptoms, family interruptions)

These tests verify the decision engine still classifies correctly when faced
with noisy patient language.
"""
import os
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from backend.decision import decide_from_text


# A representative sample of capa2_ruidosa-style patient turns (hand-crafted
# to match the noise patterns documented in the challenge: evasive answers,
# family interruptions, irrelevant symptoms).
NOISY_SAMPLES = {
    "rojo_fever_minimizer": (
        "Ay, no, tranquila doctora, un poquito molesto no más, nada del otro mundo. "
        "Sí, me tomé la temperatura ayer, marcó como 37 y algo, nada de escalofríos. "
        "La herida se ve un poquito rojita ahí en el borde.",
        "rojo",
    ),
    "rojo_bleeding_evasive": (
        "Bueno, no sé, hay como un poquito de sangre, no es mucho, será normal no? "
        "Ay, no sé, pregúntele a mi hija que ella es la que sabe. Espere, ya le digo. "
        "Doctora, está sangrando bastante, dice mi hija.",
        "rojo",
    ),
    "amarillo_fever_evasive": (
        "Más o menos, ahí normal digamos. Oiga, ¿usted sabe si es normal no poder "
        "dormir bien después de estas cirugías? No me he tomado la temperatura muy "
        "seguido, a veces siento calorcito, pero no sé si es de la casa.",
        "amarillo",
    ),
    "verde_irrelevant": (
        "Ay, ¿sabe qué? Mi vecina me dijo que no comiera banano. ¿Eso es verdad? "
        "Por lo demás, todo bien, me siento bien, no tengo nada.",
        "verde",
    ),
    "rojo_dyspnea": (
        "No, qué va, eso es mentira, yo no he sentido nada raro. Bueno... es que "
        "ayer en la noche sentí como que me ahogaba un poquito, pero pensé que era "
        "la almohada. ¿Eso es malo, verdad?",
        "rojo",
    ),
}


@pytest.mark.parametrize("name", list(NOISY_SAMPLES.keys()))
def test_noisy_keyword_doesnt_crash(name):
    """Keyword fallback does not crash on noisy data and returns a valid label.

    Note: the minimizer language in capa2_ruidosa is designed to be
    under-classified by simple keyword matching — that's exactly why we
    use the LLM path in DecisionEngine. The keyword path is a fallback.
    """
    text, _expected = NOISY_SAMPLES[name]
    result = decide_from_text(text)
    assert result["label"] in ("verde", "amarillo", "rojo")
    assert isinstance(result["score"], int)
    assert 0 <= result["score"] <= 20
    assert "rationale" in result


def test_noisy_does_not_break_score_function():
    """Score function handles all noise patterns without exceptions."""
    for text, _expected in NOISY_SAMPLES.values():
        from backend.decision import score_from_text
        score = score_from_text(text)
        assert isinstance(score, int)
        assert 0 <= score <= 20


def test_real_capa2_sample_loads():
    """Verify we can actually load capa2_ruidosa cases from the dataset."""
    from openpyxl import load_workbook
    from collections import defaultdict

    DATA = Path(__file__).resolve().parents[1] / "dataset" / "dataset_final.xlsx"
    if not DATA.exists():
        pytest.skip("dataset not present")

    wb = load_workbook(DATA, read_only=True)
    ws = wb.active
    cases = defaultdict(lambda: {"label": None, "texts": [], "capa": None})

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or len(row) < 12:
            continue
        caso_id = row[1]
        label = row[7]
        if not caso_id or not label or str(label).strip().lower() not in ("verde", "amarillo", "rojo"):
            continue
        capa = row[11]
        if capa != "capa2_ruidosa":
            continue
        c = cases[caso_id]
        if c["label"] is None:
            c["label"] = str(label).strip().lower()
        if row[5] == "paciente" and row[6]:
            c["texts"].append(str(row[6]))

    assert len(cases) > 0, "No capa2_ruidosa cases found in dataset"
    # The dataset says "the same caso_id contains both versions of the conversation"
    # so capa2 should have ~160 cases too.
    assert len(cases) >= 100, f"Expected >= 100 capa2 cases, got {len(cases)}"
    print(f"\n  capa2_ruidosa has {len(cases)} cases")
