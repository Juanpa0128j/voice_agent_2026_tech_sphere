"""Tests for clinical decision logic and severity classification."""
import pytest
from backend.decision import (
    decide_from_text,
    score_from_text,
    severity_from_score,
    STRUCTURED_DECISION_PROMPT,
)


def test_decide_from_text_no_symptoms():
    result = decide_from_text("Hola, estoy bien")
    assert result["label"] == "verde"
    assert result["alert"] is False


def test_decide_from_text_fever():
    result = decide_from_text("Tengo fiebre de 39 grados")
    assert result["label"] in ("amarillo", "rojo")
    assert result["alert"] is True


def test_decide_from_text_bleeding():
    result = decide_from_text("Estoy sangrando mucho de la herida")
    assert result["label"] == "rojo"
    assert result["alert"] is True


def test_decide_from_text_difficulty_breathing():
    result = decide_from_text("No puedo respirar bien")
    assert result["label"] == "rojo"
    assert result["alert"] is True


def test_decide_calibrated_against_dataset():
    from openpyxl import load_workbook
    wb = load_workbook("dataset/trayectorias_postop_silver.xlsx", read_only=True)
    ws = wb.active
    rojo_correct = 0
    rojo_total = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i >= 20:
            break
        if row and len(row) > 0 and row[0]:
            text = " ".join(str(c) for c in row[1:8] if c)
            label = row[0] if row[0] else "verde"
            if "rojo" in str(label).lower():
                rojo_total += 1
                result = decide_from_text(text)
                if result["label"] == "rojo":
                    rojo_correct += 1
    if rojo_total > 0:
        recall = rojo_correct / rojo_total
        assert recall >= 0.5, f"Recall of rojo: {recall}"


def test_score_from_text_keywords():
    score = score_from_text("Tengo fiebre alta y sangrado")
    assert score >= 5


def test_severity_from_score_thresholds():
    assert severity_from_score(0) == "verde"
    assert severity_from_score(2) == "verde"
    assert severity_from_score(3) == "amarillo"
    assert severity_from_score(6) == "rojo"
