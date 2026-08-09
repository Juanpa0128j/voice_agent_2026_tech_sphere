"""Patient context loading and prompt injection.

Reads patient profiles and post-operative trajectories from the vendored
xlsx dataset and produces a `PatientContext` plus a natural-language
prompt fragment intended for injection into the LLM system prompt.
"""
from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

import openpyxl


DATASET_DIR = Path(__file__).resolve().parent.parent / "dataset"

CLINICAL_FILE = DATASET_DIR / "perfiles_clinicos_pacientes_silver_contest.xlsx"
PATIENT_INFO_FILE = DATASET_DIR / "perfiles_pacientes_co.xlsx"
TRAJECTORY_FILE = DATASET_DIR / "trayectorias_postop_silver.xlsx"


@dataclass
class PatientContext:
    paciente_id: str
    nombre: str
    procedimiento: str
    dia_postoperatorio: int
    comorbilidades: List[str]
    eps: str


def _dataset_id_from_public(pid: str) -> str:
    """Map a public ID like 'P001' to the dataset's internal id 'pac_42_00000'.

    The vendored dataset uses `pac_42_00000` style ids; tests use the
    friendlier `P001` form. Unknown shapes pass through unchanged so that
    callers can also query the dataset directly.
    """
    if not pid:
        return ""
    s = pid.strip()
    if not s:
        return ""
    if s.startswith("pac_"):
        return s
    if len(s) > 1 and s[0].upper() == "P" and s[1:].isdigit():
        n = int(s[1:])
        if n >= 1:
            return f"pac_42_{n - 1:05d}"
    return s


def _parse_comorbilidades(raw) -> List[str]:
    if raw is None or raw == "":
        return []
    if isinstance(raw, list):
        return [str(x) for x in raw]
    s = str(raw).strip()
    if not s or s == "[]":
        return []
    try:
        data = json.loads(s)
    except (ValueError, TypeError):
        return [s]
    if isinstance(data, list):
        return [str(x) for x in data]
    return [str(data)]


def _read_sheet(path: Path):
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []
    header = rows[0]
    return [dict(zip(header, r)) for r in rows[1:]]


def _index_by(rows, key: str):
    out = {}
    for r in rows:
        k = r.get(key)
        if k is not None:
            out[k] = r
    return out


def _latest_trajectory_by_patient():
    """Return {paciente_id: trajectory_row} selecting the row with max dia_postop."""
    out = {}
    for r in _read_sheet(TRAJECTORY_FILE):
        pid = r.get("paciente_id")
        if not pid:
            continue
        dia = r.get("dia_postop")
        prev = out.get(pid)
        if prev is None:
            out[pid] = r
            continue
        prev_dia = prev.get("dia_postop")
        if dia is not None and (prev_dia is None or dia > prev_dia):
            out[pid] = r
    return out


def load_patient_context(paciente_id: str) -> Optional[PatientContext]:
    """Load a patient's profile plus their most recent post-op trajectory.

    Returns None if the patient id is not present in the clinical dataset.
    """
    if not paciente_id or not paciente_id.strip():
        return None

    internal_id = _dataset_id_from_public(paciente_id)

    clinical = _index_by(_read_sheet(CLINICAL_FILE), "paciente_id").get(internal_id)
    if clinical is None:
        return None

    info = _index_by(_read_sheet(PATIENT_INFO_FILE), "paciente_id").get(internal_id) or {}
    trajectory = _latest_trajectory_by_patient().get(internal_id) or {}

    comorbilidades = _parse_comorbilidades(clinical.get("comorbilidades"))
    dia = trajectory.get("dia_postop")
    if dia is None:
        dia = 0

    return PatientContext(
        paciente_id=paciente_id,
        nombre=info.get("nombre_completo") or "",
        procedimiento=clinical.get("procedimiento") or "",
        dia_postoperatorio=int(dia),
        comorbilidades=comorbilidades,
        eps=info.get("eps") or "",
    )


def build_context_prompt(ctx: PatientContext) -> str:
    """Build a natural-language prompt fragment for the LLM system prompt."""
    comorb = ctx.comorbilidades if ctx.comorbilidades else "Sin comorbilidades registradas"
    return (
        f"Paciente: {ctx.nombre} (ID: {ctx.paciente_id}).\n"
        f"Procedimiento: {ctx.procedimiento}.\n"
        f"Día postoperatorio: {ctx.dia_postoperatorio}.\n"
        f"Comorbilidades: {comorb}.\n"
        f"EPS: {ctx.eps}."
    )
