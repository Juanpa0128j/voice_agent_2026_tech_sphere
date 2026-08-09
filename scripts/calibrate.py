"""Calibrate decision thresholds against the dataset ground truth.

Reads dataset/dataset_final.xlsx (3991 turn rows from 160 cases), aggregates
all patient turns per case, runs decision.decide_from_text() on the concatenated
text, and compares with label_ground_truth.

Prints:
  - Total cases evaluated
  - Accuracy overall
  - Recall of "rojo" (the critical metric for clinical safety)
  - Confusion matrix

Also tests capa2_ruidosa (the noisy version) for robustness.
"""
import sys
from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from backend.decision import decide_from_text

DATA = Path(__file__).resolve().parents[1] / "dataset" / "dataset_final.xlsx"

# Map dataset ground truth labels (these appear consistent across rows of a case)
# Per the dataset: caso_id = "caso_" + trayectoria_id, and label is constant per caso_id
COL_LABEL = 7
COL_CASO = 1
COL_HABLANTE = 5
COL_TEXTO = 6
COL_CAPA = 11


def load_cases(layer: str = "capa1_limpia"):
    """Return {caso_id: {"label": str, "text": str, "capa": str}}"""
    wb = load_workbook(DATA, read_only=True)
    ws = wb.active
    cases: dict = defaultdict(lambda: {"label": None, "text_parts": [], "capa": None})
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or len(row) < 8:
            continue
        caso_id = row[COL_CASO]
        label = row[COL_LABEL]
        hablante = row[COL_HABLANTE]
        texto = row[COL_TEXTO]
        capa = row[COL_CAPA] if len(row) > COL_CAPA else None
        if not caso_id or not label:
            continue
        if capa and capa != layer:
            continue
        case = cases[caso_id]
        if case["label"] is None:
            case["label"] = str(label).strip().lower()
            case["capa"] = capa
        if hablante == "paciente" and texto:
            case["text_parts"].append(str(texto))
    return {
        cid: {"label": c["label"], "text": " ".join(c["text_parts"]), "capa": c["capa"]}
        for cid, c in cases.items()
        if c["label"] in ("verde", "amarillo", "rojo")
    }


def evaluate(cases, title: str):
    confusion = Counter()
    correct = 0
    rojo_total = 0
    rojo_correct = 0
    amarillo_total = 0
    amarillo_correct = 0
    verde_total = 0
    verde_correct = 0

    for cid, c in cases.items():
        text = c["text"]
        if not text.strip():
            continue
        gt = c["label"]
        result = decide_from_text(text)
        pred = result["label"]
        confusion[(gt, pred)] += 1
        if pred == gt:
            correct += 1
        if gt == "rojo":
            rojo_total += 1
            if pred == "rojo":
                rojo_correct += 1
        elif gt == "amarillo":
            amarillo_total += 1
            if pred == "amarillo":
                amarillo_correct += 1
        elif gt == "verde":
            verde_total += 1
            if pred == "verde":
                verde_correct += 1

    total = sum(confusion.values())
    print(f"\n=== {title} ===")
    print(f"Total cases: {total}")
    if total:
        print(f"Overall accuracy: {correct/total:.2%} ({correct}/{total})")
    if rojo_total:
        print(f"Recall 'rojo' (critical): {rojo_correct/rojo_total:.2%} ({rojo_correct}/{rojo_total})")
    if amarillo_total:
        print(f"Recall 'amarillo': {amarillo_correct/amarillo_total:.2%} ({amarillo_correct}/{amarillo_total})")
    if verde_total:
        print(f"Recall 'verde': {verde_correct/verde_total:.2%} ({verde_correct}/{verde_total})")

    print(f"\nConfusion matrix (rows=ground truth, cols=predicted):")
    print(f"{'':12s} {'verde':>8s} {'amarillo':>10s} {'rojo':>8s}")
    for gt in ("verde", "amarillo", "rojo"):
        row = [confusion.get((gt, p), 0) for p in ("verde", "amarillo", "rojo")]
        print(f"{gt:12s} {row[0]:>8d} {row[1]:>10d} {row[2]:>8d}")

    return {
        "total": total,
        "accuracy": correct / total if total else 0,
        "recall_rojo": rojo_correct / rojo_total if rojo_total else 0,
    }


def main() -> None:
    clean = load_cases("capa1_limpia")
    noisy = load_cases("capa2_ruidosa")
    res_clean = evaluate(clean, "Capa 1 (clean) - ground truth")
    res_noisy = evaluate(noisy, "Capa 2 (noisy) - robustness test")

    # Persist results for the README
    import json
    out = {
        "capa1_limpia": res_clean,
        "capa2_ruidosa": res_noisy,
    }
    Path("backend/calibration_results.json").write_text(json.dumps(out, indent=2))
    print("\nResults saved to backend/calibration_results.json")


if __name__ == "__main__":
    main()
